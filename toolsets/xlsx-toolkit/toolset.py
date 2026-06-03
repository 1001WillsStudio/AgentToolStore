"""
xlsx‑toolkit — Read, write, and convert Excel spreadsheets.
================================================================

Provides four functions for working with .xlsx files without needing
to write throw‑away scripts for every spreadsheet task.
"""

from pathlib import Path

try:
    from toolstore.toolset import tool
except ImportError:
    def tool(fn):
        return fn  # no‑op when toolstore package not installed


# ── xlsx_read ──────────────────────────────────────────────────────────

@tool
def xlsx_read(*, filepath: str, sheet: str = "",
              max_rows: int = 0) -> dict:
    """Read an Excel sheet and return rows as a JSON array.

    Args:
        filepath: Absolute or relative path to the .xlsx file.
        sheet:     Sheet name (default: first sheet).
        max_rows:  Limit rows read (0 = all).

    Returns:
        dict with keys:
          sheet   — name of the sheet that was read
          columns — list of column headers
          rows    — list of dicts, one per data row
          count   — number of data rows returned
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"error": "openpyxl not installed — run: pip install openpyxl"}

    p = Path(filepath).expanduser().resolve()
    if not p.exists():
        return {"error": f"File not found: {p}"}

    try:
        wb = load_workbook(p, read_only=True, data_only=True)
    except Exception as exc:
        return {"error": f"Cannot open workbook: {exc}"}

    ws = wb[sheet] if sheet else wb.active
    if ws is None:
        return {"error": f"Sheet '{sheet or '(default)'}' not found"}

    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [str(h) if h is not None else f"col_{i}"
                   for i, h in enumerate(next(rows_iter))]
    except StopIteration:
        return {"sheet": ws.title, "columns": [], "rows": [], "count": 0}

    rows = []
    for row in rows_iter:
        row_dict = {headers[i]: (None if v is None else v)
                    for i, v in enumerate(row) if i < len(headers)}
        for h in headers[len(row):]:
            row_dict[h] = None
        rows.append(row_dict)
        if max_rows and len(rows) >= max_rows:
            break

    wb.close()
    return {"sheet": ws.title, "columns": headers, "rows": rows, "count": len(rows)}


# ── xlsx_sheets ────────────────────────────────────────────────────────

@tool
def xlsx_sheets(*, filepath: str) -> dict:
    """List all sheet names in an Excel file with row/column counts.

    Args:
        filepath: Path to the .xlsx file.

    Returns:
        dict with keys:
          filename — base name of the file
          sheets   — list of {name, rows, cols}
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"error": "openpyxl not installed — run: pip install openpyxl"}

    p = Path(filepath).expanduser().resolve()
    if not p.exists():
        return {"error": f"File not found: {p}"}

    try:
        wb = load_workbook(p, read_only=True, data_only=True)
    except Exception as exc:
        return {"error": f"Cannot open workbook: {exc}"}

    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        sheets.append({"name": name, "rows": ws.max_row or 0, "cols": ws.max_column or 0})

    wb.close()
    return {"filename": p.name, "sheets": sheets}


# ── xlsx_to_csv ────────────────────────────────────────────────────────

@tool
def xlsx_to_csv(*, filepath: str, sheet: str = "",
                delimiter: str = ",") -> dict:
    """Convert an Excel sheet to CSV text.

    Args:
        filepath:  Path to the .xlsx file.
        sheet:     Sheet name (default: first sheet).
        delimiter: Field delimiter (default comma).

    Returns:
        dict with key "csv" containing the full CSV string.
    """
    result = xlsx_read(filepath=filepath, sheet=sheet)
    if "error" in result:
        return result

    columns = result["columns"]
    lines = [delimiter.join(f'"{c}"' if delimiter in str(c) else str(c)
                            for c in columns)]
    for row in result["rows"]:
        lines.append(delimiter.join(
            f'"{v}"' if v is not None and delimiter in str(v) else (str(v) if v is not None else "")
            for v in (row.get(c, "") for c in columns)
        ))
    return {"csv": "\n".join(lines), "sheet": result["sheet"],
            "rows": result["count"], "columns": columns}


# ── xlsx_create ────────────────────────────────────────────────────────

@tool
def xlsx_create(*, filepath: str, sheet: str = "Sheet1",
                columns: list, rows: list) -> dict:
    """Create a new .xlsx file from provided data.

    Args:
        filepath: Where to write the new .xlsx file.
        sheet:    Sheet name.
        columns:  List of column header strings.
        rows:     List of dicts, each keyed by column name.

    Returns:
        dict with "written" (absolute path) or "error".
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return {"error": "openpyxl not installed — run: pip install openpyxl"}

    p = Path(filepath).expanduser().resolve()
    p.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align

    for r, row_data in enumerate(rows, 2):
        for c, col_name in enumerate(columns, 1):
            value = row_data.get(col_name) if isinstance(row_data, dict) else row_data[c-1] if isinstance(row_data, list) else None
            ws.cell(row=r, column=c, value=value)

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    wb.save(p)
    return {"written": str(p), "sheet": sheet,
            "columns": len(columns), "rows": len(rows)}
